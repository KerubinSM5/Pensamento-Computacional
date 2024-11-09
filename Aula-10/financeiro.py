from openpyxl import Workbook
from util.db import SQL
from util import converter as cv


class Financeiro:

    def __init__(self):
        self.arquivo = 'financeiro.xlsx'
        self.idt_projeto = 0
        self.sql = SQL(esquema='bd_planejamento')

        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Planilha Financeira"
        self.ws['A1'] = "Projeto:"
        self.ws['A2'] = "Colaborador"
        self.ws['B2'] = "Início"
        self.ws['C2'] = "Término"
        self.ws['D2'] = "Número de Dias"
        self.ws['E2'] = "Função"
        self.ws['F2'] = "Valor da Diária"
        self.ws['G2'] = "Total"
        self.ws.column_dimensions['A'].width = 30
        self.ws.column_dimensions['B'].width = 15
        self.ws.column_dimensions['C'].width = 15
        self.ws.column_dimensions['D'].width = 15
        self.ws.column_dimensions['E'].width = 25
        self.ws.column_dimensions['F'].width = 15

    def buscar_projeto(self, idt_projeto):
        cmd = "SELECT nme_projeto FROM tb_projeto WHERE idt_projeto = %s;"
        projeto = self.sql.get_string(cmd, [idt_projeto])
        return projeto

    def buscar_colaboradores(self, idt_projeto):
        cmd = """
          SELECT nme_colaborador, dta_ini_alocacao, dta_fim_alocacao,
           datediff(dta_fim_alocacao, dta_ini_alocacao) AS dias,
           nme_funcao, vlr_dia_funcao,
           (datediff(dta_fim_alocacao, dta_ini_alocacao) * vlr_dia_funcao) as tot_col
          FROM tb_colaborador JOIN ta_alocacao ON idt_colaborador = cod_colaborador
          JOIN tb_funcao ON idt_funcao = cod_funcao
          WHERE cod_projeto = %s
          ORDER BY nme_colaborador;
       """
        lista = self.sql.get_list(cmd, [idt_projeto])
        return lista

    def gerar_planilha(self):
        # Buscar o Projeto
        projeto = self.buscar_projeto(self.idt_projeto)
        self.ws['B1'] = projeto

        # Buscar a parte financeira
        total_geral = 0.0
        linha = 3
        lista = self.buscar_colaboradores(self.idt_projeto)
        for obj in lista:
            self.ws.cell(row=linha, column=1, value=obj['nme_colaborador'])
            self.ws.cell(row=linha, column=2, value=cv.mysql_to_bra(obj['dta_ini_alocacao']))
            self.ws.cell(row=linha, column=3, value=cv.mysql_to_bra(obj['dta_fim_alocacao']))
            self.ws.cell(row=linha, column=4, value=obj['dias'])
            self.ws.cell(row=linha, column=5, value=obj['nme_funcao'])
            self.ws.cell(row=linha, column=6, value=obj['vlr_dia_funcao'])
            self.ws.cell(row=linha, column=7, value=obj['tot_col'])
            total_geral += float(obj['tot_col'])
            linha += 1

        self.ws.cell(row=linha, column=6, value="Total Geral")
        self.ws.cell(row=linha, column=7, value=total_geral)

    def salvar(self):
        self.wb.save(self.arquivo)
        print("Planilha salva ...")


if __name__ == '__main__':
    financeiro = Financeiro()
    financeiro.idt_projeto = 2
    financeiro.gerar_planilha()
    financeiro.salvar()
